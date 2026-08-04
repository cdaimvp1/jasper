"""Regression tests for attachment_extract.py (task #29, 2026-08-01):
deterministic text-layer/cell-value extraction from real attachment files,
so value/asks extraction can read what's actually attached, not just the
email body around it."""
from __future__ import annotations

import zipfile
from pathlib import Path

import openpyxl
import pytest

import attachment_extract as ae

_DOCX_CONTENT_TYPES = (
    b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
    b'<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
    b'<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
    b'<Override PartName="/word/document.xml" '
    b'ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>'
    b"</Types>"
)
_DOCX_RELS = (
    b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
    b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
    b'<Relationship Id="rId1" '
    b'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" '
    b'Target="word/document.xml"/></Relationships>'
)


def _minimal_docx_bytes(paragraphs: list[str]) -> bytes:
    """A real, valid, minimal .docx (a zip containing a real
    word/document.xml with one <w:p> per paragraph, each holding one
    <w:t> run) - built by hand the same way test_extract_pdf_text's own
    fixture builds a real minimal PDF, not a mock."""
    ns = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
    body_paras = "".join(
        f'<w:p><w:r><w:t>{p}</w:t></w:r></w:p>' for p in paragraphs
    )
    document_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<w:document xmlns:w="{ns}"><w:body>{body_paras}</w:body></w:document>'
    ).encode()

    import io
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        zf.writestr("[Content_Types].xml", _DOCX_CONTENT_TYPES)
        zf.writestr("_rels/.rels", _DOCX_RELS)
        zf.writestr("word/document.xml", document_xml)
    return buf.getvalue()


def _minimal_pdf_bytes(text: str) -> bytes:
    """A real, valid, minimal one-page PDF containing exactly `text` as its
    visible content - a correct xref table (byte offsets computed, not
    guessed) so pypdf parses it via the normal path, not an error-recovery
    fallback. Verified round-trips through pypdf before use here."""
    stream_content = f"BT /F1 18 Tf 10 100 Td ({text}) Tj ET".encode()
    parts: list[bytes] = [b"%PDF-1.4\n"]
    offsets: list[int] = []

    def add(obj_bytes: bytes) -> None:
        offsets.append(sum(len(p) for p in parts))
        parts.append(obj_bytes)

    add(b"1 0 obj<</Type/Catalog/Pages 2 0 R>>endobj\n")
    add(b"2 0 obj<</Type/Pages/Kids[3 0 R]/Count 1>>endobj\n")
    add(b"3 0 obj<</Type/Page/Parent 2 0 R/Resources<</Font<</F1 4 0 R>>>>/MediaBox[0 0 300 144]/Contents 5 0 R>>endobj\n")
    add(b"4 0 obj<</Type/Font/Subtype/Type1/BaseFont/Helvetica>>endobj\n")
    add(b"5 0 obj<</Length " + str(len(stream_content)).encode() + b">>\nstream\n" + stream_content + b"\nendstream\nendobj\n")

    xref_offset = sum(len(p) for p in parts)
    xref = b"xref\n0 6\n0000000000 65535 f \n"
    for off in offsets:
        xref += ("%010d 00000 n \n" % off).encode()
    trailer = b"trailer\n<</Size 6/Root 1 0 R>>\nstartxref\n" + str(xref_offset).encode() + b"\n%%EOF"

    return b"".join(parts) + xref + trailer


# --- PDF extraction -------------------------------------------------------

def test_extract_pdf_text_finds_real_text(tmp_path):
    path = tmp_path / "order_form.pdf"
    path.write_bytes(_minimal_pdf_bytes("Total contract value: $50,000,000"))

    assert ae.extract_pdf_text(path) == "Total contract value: $50,000,000"


def test_extract_pdf_text_fails_open_on_corrupt_file(tmp_path):
    """A malformed/encrypted/non-PDF file must never raise - one bad
    attachment must not break the ingest batch it arrived in."""
    path = tmp_path / "corrupt.pdf"
    path.write_bytes(b"this is not a real pdf at all")

    assert ae.extract_pdf_text(path) == ""


def test_extract_pdf_text_fails_open_on_missing_file(tmp_path):
    assert ae.extract_pdf_text(tmp_path / "does_not_exist.pdf") == ""


# --- XLSX extraction -------------------------------------------------------

def test_extract_xlsx_text_finds_real_cell_values(tmp_path):
    path = tmp_path / "pricing.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Line item"
    ws["B1"] = "Amount"
    ws["A2"] = "Subscription fee"
    ws["B2"] = 53702143
    wb.save(path)

    text = ae.extract_xlsx_text(path)

    assert "Subscription fee" in text
    assert "53702143" in text


def test_extract_xlsx_text_reads_multiple_worksheets(tmp_path):
    path = tmp_path / "multi.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"
    wb.active["A1"] = "first sheet value"
    wb.create_sheet("Sheet2")["A1"] = "second sheet value"
    wb.save(path)

    text = ae.extract_xlsx_text(path)

    assert "first sheet value" in text
    assert "second sheet value" in text


def test_extract_xlsx_text_fails_open_on_corrupt_file(tmp_path):
    path = tmp_path / "corrupt.xlsx"
    path.write_bytes(b"not a real xlsx file")

    assert ae.extract_xlsx_text(path) == ""


def test_extract_xlsx_text_fails_open_on_missing_file(tmp_path):
    assert ae.extract_xlsx_text(tmp_path / "does_not_exist.xlsx") == ""


# --- DOCX extraction (enhancement idea panel #7/worker capability #6) -----

def test_extract_docx_text_finds_real_paragraph_text(tmp_path):
    path = tmp_path / "contract.docx"
    path.write_bytes(_minimal_docx_bytes(["Total contract value: $50,000,000"]))

    assert ae.extract_docx_text(path) == "Total contract value: $50,000,000"


def test_extract_docx_text_joins_multiple_paragraphs_with_newline(tmp_path):
    path = tmp_path / "multi.docx"
    path.write_bytes(_minimal_docx_bytes(["First paragraph", "Second paragraph"]))

    assert ae.extract_docx_text(path) == "First paragraph\nSecond paragraph"


def test_extract_docx_text_fails_open_on_corrupt_file(tmp_path):
    path = tmp_path / "corrupt.docx"
    path.write_bytes(b"not a real docx file")

    assert ae.extract_docx_text(path) == ""


def test_extract_docx_text_fails_open_on_missing_file(tmp_path):
    assert ae.extract_docx_text(tmp_path / "does_not_exist.docx") == ""


def test_extract_docx_text_empty_body_returns_empty_string(tmp_path):
    path = tmp_path / "empty.docx"
    path.write_bytes(_minimal_docx_bytes([]))

    assert ae.extract_docx_text(path) == ""


# --- dispatch ---------------------------------------------------------------

def test_extract_text_dispatches_pdf(tmp_path):
    path = tmp_path / "doc.pdf"
    path.write_bytes(_minimal_pdf_bytes("dispatch check"))
    assert ae.extract_text(path) == "dispatch check"


def test_extract_text_dispatches_xlsx(tmp_path):
    path = tmp_path / "doc.xlsx"
    wb = openpyxl.Workbook()
    wb.active["A1"] = "dispatch check"
    wb.save(path)
    assert "dispatch check" in ae.extract_text(path)


def test_extract_text_dispatch_is_case_insensitive(tmp_path):
    path = tmp_path / "doc.PDF"
    path.write_bytes(_minimal_pdf_bytes("uppercase extension"))
    assert ae.extract_text(path) == "uppercase extension"


def test_extract_text_dispatches_docx(tmp_path):
    path = tmp_path / "doc.docx"
    path.write_bytes(_minimal_docx_bytes(["dispatch check"]))
    assert ae.extract_text(path) == "dispatch check"


def test_extract_text_unknown_extension_returns_empty_not_a_guess(tmp_path):
    """A real, honest gap for any type with no registered extractor. Must
    return empty, never something that looks like real content."""
    path = tmp_path / "doc.pptx"
    path.write_bytes(b"some pptx bytes")
    assert ae.extract_text(path) == ""
