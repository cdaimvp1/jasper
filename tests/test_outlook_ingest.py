"""Regression tests for task #43: outlook_com_ingest.py used to truncate every
body to 500 chars and never persist Outlook's own EntryID - both fixed by
staging the full plain-text/HTML body to files (same pattern attachments
already use) and adding a real entry_id column to raw_items.

No live Outlook involved - outlook_scan.ps1 is a COM wrapper that can't run
in CI/off-Windows-Outlook anyway, so subprocess.run is monkeypatched to return
JSON-lines shaped exactly like the real script's new output, and item_staged_dir
points at a real tmp_path directory containing real body.txt/body.html files -
this is the same "shape the mock data 1:1 with the real emitter" discipline
used for every other ingestion test this session.
"""
from __future__ import annotations

import json
import subprocess
import sys
from pathlib import Path

BODY = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BODY / "ingest"))
import outlook_com_ingest as oci  # noqa: E402


class _FakeCompletedProcess:
    def __init__(self, stdout: str, returncode: int = 0, stderr: str = ""):
        self.stdout = stdout
        self.returncode = returncode
        self.stderr = stderr


def _stage_item(tmp_path, name: str, entry_id: str, body_text: str, body_html: str,
                 received_epoch: float = 1_800_000_000.0) -> tuple[dict, Path]:
    staged_dir = tmp_path / f"staged_{name}"
    staged_dir.mkdir()
    (staged_dir / "body.txt").write_text(body_text, encoding="utf-8")
    (staged_dir / "body.html").write_text(body_html, encoding="utf-8")
    item = {
        "conversation_id": f"conv-{name}",
        "entry_id": entry_id,
        "subject": f"Subject {name}",
        "sender": "vendor@example.com",
        "sender_name": "Vendor Example",
        "participants": ["vendor@example.com", "marc@example.com"],
        "received_epoch": received_epoch,
        "body_preview": body_text[:500],
        "attachments": [],
        "body_text_file": "body.txt",
        "body_html_file": "body.html",
        "item_staged_dir": str(staged_dir),
    }
    return item, staged_dir


def test_full_body_and_entry_id_persisted(ws_db, isolated_paths, monkeypatch, tmp_path):
    full_text = "line one\n" * 200  # far past the old 500-char truncation
    full_html = "<html><body>" + ("<p>hi</p>" * 200) + "</body></html>"
    item, staged_dir = _stage_item(tmp_path, "a", "entryid-AAA", full_text, full_html)

    def fake_run(*a, **kw):
        return _FakeCompletedProcess(json.dumps(item) + "\n")

    monkeypatch.setattr(subprocess, "run", fake_run)
    result = oci.run(folder="Careful")

    assert result["ok"] is True
    assert result["inserted"] == 1

    rows = ws_db._connect().execute("SELECT * FROM raw_items").fetchall()
    assert len(rows) == 1
    row = dict(rows[0])
    assert row["entry_id"] == "entryid-AAA"
    assert row["body_preview"] == full_text[:500]  # unchanged short-preview behavior

    ref = json.loads(row["raw_ref"])
    text_path = isolated_paths.DOCUMENTS_DIR / ref["body_text"]
    html_path = isolated_paths.DOCUMENTS_DIR / ref["body_html"]
    assert text_path.read_text(encoding="utf-8") == full_text
    assert html_path.read_text(encoding="utf-8") == full_html


def test_staged_dir_cleaned_up_after_absorb(ws_db, isolated_paths, monkeypatch, tmp_path):
    item, staged_dir = _stage_item(tmp_path, "b", "entryid-BBB", "body", "<p>body</p>")

    def fake_run(*a, **kw):
        return _FakeCompletedProcess(json.dumps(item) + "\n")

    monkeypatch.setattr(subprocess, "run", fake_run)
    oci.run(folder="Careful")

    assert not staged_dir.exists()


def test_duplicate_item_still_cleans_staged_dir(ws_db, isolated_paths, monkeypatch, tmp_path):
    """Same dedupe_key twice (a replayed cursor) - the second insert is a
    no-op duplicate, but its staging folder must still be reclaimed, exactly
    the same guarantee run() already made for attachments before this change."""
    item1, dir1 = _stage_item(tmp_path, "c1", "entryid-CCC1", "body", "<p>b</p>",
                               received_epoch=1_800_000_100.0)
    item2, dir2 = _stage_item(tmp_path, "c2", "entryid-CCC1", "body", "<p>b</p>",
                               received_epoch=1_800_000_100.0)
    # Force an identical dedupe_key: same participants/day/source_ref as item1.
    item2["entry_id"] = item1["entry_id"]
    item2["conversation_id"] = item1["conversation_id"]
    item2["participants"] = item1["participants"]

    calls = iter([
        _FakeCompletedProcess(json.dumps(item1) + "\n"),
        _FakeCompletedProcess(json.dumps(item2) + "\n"),
    ])
    monkeypatch.setattr(subprocess, "run", lambda *a, **kw: next(calls))

    r1 = oci.run(folder="Careful")
    r2 = oci.run(folder="Careful")

    assert r1["inserted"] == 1
    assert r2["duplicates"] == 1
    assert not dir1.exists()
    assert not dir2.exists()


def test_absorb_body_handles_missing_files_gracefully(ws_db, isolated_paths, tmp_path):
    staged_dir = tmp_path / "staged_missing"
    staged_dir.mkdir()  # no body.txt / body.html actually written inside

    ref = oci._absorb_body(row_id=999, item_staged_dir=str(staged_dir),
                            text_file="body.txt", html_file="body.html")
    assert ref is None  # neither file existed - nothing to point at, no crash


def test_absorb_body_returns_none_when_no_staged_dir(ws_db, isolated_paths):
    assert oci._absorb_body(row_id=1, item_staged_dir=None,
                             text_file="body.txt", html_file="body.html") is None


# --- attachment dedup + extraction (task #29, 2026-08-01) -----------------

def _staged_attachment(staging_dir: Path, filename: str, content: bytes) -> dict:
    path = staging_dir / filename
    path.write_bytes(content)
    return {"filename": filename, "staged_path": str(path), "size_bytes": len(content)}


def test_absorb_attachments_extracts_text_from_a_real_xlsx(ws_db, isolated_paths, tmp_path):
    import openpyxl
    staging_dir = tmp_path / "att_staging_1"
    staging_dir.mkdir()
    xlsx_path = staging_dir / "pricing.xlsx"
    wb = openpyxl.Workbook()
    wb.active["A1"] = "Total contract value"
    wb.active["B1"] = 53702143
    wb.save(xlsx_path)
    staged = [{"filename": "pricing.xlsx", "staged_path": str(xlsx_path), "size_bytes": xlsx_path.stat().st_size}]

    absorbed = oci._absorb_attachments(101, staged)

    assert absorbed == 1
    rows = ws_db.list_attachments("raw_item", "101")
    assert len(rows) == 1
    assert "Total contract value" in rows[0]["extracted_text"]
    assert "53702143" in rows[0]["extracted_text"]


def test_absorb_attachments_extracts_text_from_a_real_docx(ws_db, isolated_paths, tmp_path):
    """Enhancement idea panel #7's real remaining gap, closed - a redline/
    contract .docx now gets its text extracted the same way a .pdf/.xlsx
    already did."""
    import zipfile
    staging_dir = tmp_path / "att_staging_docx"
    staging_dir.mkdir()
    docx_path = staging_dir / "contract.docx"
    ns = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
    document_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<w:document xmlns:w="{ns}"><w:body>'
        '<w:p><w:r><w:t>Total contract value: $53,702,143</w:t></w:r></w:p>'
        '</w:body></w:document>'
    ).encode()
    with zipfile.ZipFile(docx_path, "w") as zf:
        zf.writestr("[Content_Types].xml", b"<Types/>")
        zf.writestr("word/document.xml", document_xml)
    staged = [{"filename": "contract.docx", "staged_path": str(docx_path), "size_bytes": docx_path.stat().st_size}]

    absorbed = oci._absorb_attachments(107, staged)

    assert absorbed == 1
    rows = ws_db.list_attachments("raw_item", "107")
    assert len(rows) == 1
    assert "Total contract value: $53,702,143" in rows[0]["extracted_text"]


# --- backfill_docx_extracted_text (E6) -------------------------------------

def _docx_bytes(text):
    import zipfile
    import io
    ns = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
    document_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<w:document xmlns:w="{ns}"><w:body><w:p><w:r><w:t>{text}</w:t></w:r></w:p></w:body></w:document>'
    ).encode()
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        zf.writestr("[Content_Types].xml", b"<Types/>")
        zf.writestr("word/document.xml", document_xml)
    return buf.getvalue()


def test_backfill_docx_extracted_text_fills_in_real_preexisting_rows(ws_db, isolated_paths):
    a = ws_db.create_issue_with_new_id(title="A", state="active", category="other")
    doc_dir = isolated_paths.DOCUMENTS_DIR / "raw_items" / "1"
    doc_dir.mkdir(parents=True)
    (doc_dir / "contract.docx").write_bytes(_docx_bytes("real contract text $99"))
    aid = ws_db.create_attachment(
        entity_type="issue", entity_id=a, kind="upload", filename="contract.docx",
        stored_path="raw_items/1/contract.docx", content_type=None, size_bytes=10,
        sha256_hex=None, uploaded_by="marc",
    )

    result = oci.backfill_docx_extracted_text()

    assert result["updated"] == 1
    assert "real contract text $99" in ws_db.get_attachment(aid)["extracted_text"]


def test_backfill_docx_extracted_text_is_idempotent(ws_db, isolated_paths):
    a = ws_db.create_issue_with_new_id(title="A", state="active", category="other")
    doc_dir = isolated_paths.DOCUMENTS_DIR / "raw_items" / "2"
    doc_dir.mkdir(parents=True)
    (doc_dir / "contract.docx").write_bytes(_docx_bytes("text"))
    ws_db.create_attachment(
        entity_type="issue", entity_id=a, kind="upload", filename="contract.docx",
        stored_path="raw_items/2/contract.docx", content_type=None, size_bytes=10,
        sha256_hex=None, uploaded_by="marc",
    )

    first = oci.backfill_docx_extracted_text()
    second = oci.backfill_docx_extracted_text()

    assert first["updated"] == 1
    assert second["updated"] == 0  # already filled in - nothing left to do


def test_backfill_docx_extracted_text_skips_missing_file(ws_db, isolated_paths):
    a = ws_db.create_issue_with_new_id(title="A", state="active", category="other")
    ws_db.create_attachment(
        entity_type="issue", entity_id=a, kind="upload", filename="ghost.docx",
        stored_path="raw_items/3/ghost.docx", content_type=None, size_bytes=10,
        sha256_hex=None, uploaded_by="marc",
    )

    result = oci.backfill_docx_extracted_text()

    assert result["updated"] == 0
    assert result["skipped_missing_file"] == 1


def test_absorb_attachments_dedupes_byte_identical_files_across_raw_items(ws_db, isolated_paths, tmp_path):
    """The exact real gap: the same real document forwarded across several
    emails used to get copied and text-extracted once per email. A second
    raw_item with a byte-identical attachment must reuse the first's stored
    file and extracted text, not duplicate either."""
    import openpyxl
    content_dir = tmp_path / "att_staging_2"
    content_dir.mkdir()
    xlsx_path_1 = content_dir / "order_form_v1.xlsx"
    wb = openpyxl.Workbook()
    wb.active["A1"] = "shared content"
    wb.save(xlsx_path_1)
    original_bytes = xlsx_path_1.read_bytes()

    staged_1 = [_staged_attachment(content_dir, "order_form_copy1.xlsx", original_bytes)]
    absorbed_1 = oci._absorb_attachments(102, staged_1)
    assert absorbed_1 == 1
    first_row = ws_db.list_attachments("raw_item", "102")[0]
    assert "shared content" in first_row["extracted_text"]

    # Second raw_item, byte-identical file content, different filename (a
    # real-world forward often renames slightly) - same hash either way.
    staged_2 = [_staged_attachment(content_dir, "order_form_copy2.xlsx", original_bytes)]
    absorbed_2 = oci._absorb_attachments(103, staged_2)
    assert absorbed_2 == 1
    second_row = ws_db.list_attachments("raw_item", "103")[0]

    assert second_row["stored_path"] == first_row["stored_path"], "must reuse the SAME stored file, not copy again"
    assert second_row["extracted_text"] == first_row["extracted_text"], "must reuse the cached extraction, not re-run it"

    # Only ONE physical copy of the document ever landed in the document library.
    doc_files = list((isolated_paths.DOCUMENTS_RAW_ITEMS_DIR / "102").glob("*.xlsx"))
    assert len(doc_files) == 1
    assert not (isolated_paths.DOCUMENTS_RAW_ITEMS_DIR / "103").exists(), "the second raw_item's own dir was never created - nothing new to store there"

    # Design doc Section 12.5: the same real hash match that drives dedup
    # is also the live producer for artifact_lineages - two byte-identical
    # attachments across two raw_items must land in ONE lineage with TWO
    # versions, not stay invisible the way this exact case did before v2.6.
    version_1 = ws_db.find_artifact_version_by_attachment(first_row["id"])
    version_2 = ws_db.find_artifact_version_by_attachment(second_row["id"])
    assert version_1 is not None and version_2 is not None
    assert version_1["lineage_id"] == version_2["lineage_id"]
    versions = ws_db.list_artifact_versions_for_lineage(version_1["lineage_id"])
    assert len(versions) == 2


def test_absorb_attachments_different_content_same_name_stores_both(ws_db, isolated_paths, tmp_path):
    """A genuinely different version (v2 of the same document) must NOT be
    deduped just because the filename matches - only a real hash match
    counts as "the same file"."""
    import openpyxl
    staging_dir = tmp_path / "att_staging_3"
    staging_dir.mkdir()

    path_v1 = staging_dir / "order_form_v1_src.xlsx"
    wb1 = openpyxl.Workbook()
    wb1.active["A1"] = "version 1 content"
    wb1.save(path_v1)
    staged_1 = [_staged_attachment(staging_dir, "order_form.xlsx", path_v1.read_bytes())]
    oci._absorb_attachments(104, staged_1)

    path_v2 = staging_dir / "order_form_v2_src.xlsx"
    wb2 = openpyxl.Workbook()
    wb2.active["A1"] = "version 2 content, genuinely different"
    wb2.save(path_v2)
    staged_2 = [_staged_attachment(staging_dir, "order_form.xlsx", path_v2.read_bytes())]
    oci._absorb_attachments(105, staged_2)

    row_v1 = ws_db.list_attachments("raw_item", "104")[0]
    row_v2 = ws_db.list_attachments("raw_item", "105")[0]
    assert row_v1["stored_path"] != row_v2["stored_path"]
    assert "version 1" in row_v1["extracted_text"]
    assert "version 2" in row_v2["extracted_text"]


def test_absorb_attachments_unsupported_extension_stores_null_extracted_text(ws_db, isolated_paths, tmp_path):
    """.pptx has no registered extractor (unlike .docx, now real - see
    attachment_extract.py's extract_docx_text)."""
    staging_dir = tmp_path / "att_staging_4"
    staging_dir.mkdir()
    staged = [_staged_attachment(staging_dir, "notes.pptx", b"some pptx bytes, not really parsed")]

    oci._absorb_attachments(106, staged)

    row = ws_db.list_attachments("raw_item", "106")[0]
    assert row["extracted_text"] is None


def test_sweep_unread_also_persists_entry_id_and_body(ws_db, isolated_paths, monkeypatch, tmp_path):
    item, staged_dir = _stage_item(tmp_path, "d", "entryid-DDD", "unread body", "<p>unread</p>")

    def fake_run(*a, **kw):
        return _FakeCompletedProcess(json.dumps(item) + "\n")

    monkeypatch.setattr(subprocess, "run", fake_run)
    result = oci.sweep_unread(folder="Careful")

    assert result["inserted"] == 1
    row = dict(ws_db._connect().execute("SELECT * FROM raw_items").fetchall()[0])
    assert row["entry_id"] == "entryid-DDD"
    assert row["raw_ref"] is not None
    assert not staged_dir.exists()


# --- task #149: stale local Outlook cache diagnostic ----------------------

def test_cold_start_diagnostic_persisted_true(ws_db, isolated_paths, monkeypatch):
    def fake_run(*a, **kw):
        return _FakeCompletedProcess("", stderr="JASPER_DIAG: outlook_was_running=False\n")

    monkeypatch.setattr(subprocess, "run", fake_run)
    result = oci.run(folder="Careful")

    assert result["outlook_cold_started"] is True
    assert ws_db.get_cursor("outlook_mail", "last_scan_outlook_cold_started") == "true"
    assert ws_db.get_cursor("outlook_mail", "consecutive_cold_starts") == "1"


def test_cold_start_diagnostic_persisted_false_resets_streak(ws_db, isolated_paths, monkeypatch):
    ws_db.set_cursor("outlook_mail", "consecutive_cold_starts", "2")

    def fake_run(*a, **kw):
        return _FakeCompletedProcess("", stderr="JASPER_DIAG: outlook_was_running=True\n")

    monkeypatch.setattr(subprocess, "run", fake_run)
    result = oci.run(folder="Careful")

    assert result["outlook_cold_started"] is False
    assert ws_db.get_cursor("outlook_mail", "consecutive_cold_starts") == "0"


def test_cold_start_streak_increments_across_runs(ws_db, isolated_paths, monkeypatch):
    def fake_run(*a, **kw):
        return _FakeCompletedProcess("", stderr="JASPER_DIAG: outlook_was_running=False\n")

    monkeypatch.setattr(subprocess, "run", fake_run)
    oci.run(folder="Careful")
    oci.run(folder="Careful")
    result = oci.run(folder="Careful")

    assert result["outlook_cold_started"] is True
    assert ws_db.get_cursor("outlook_mail", "consecutive_cold_starts") == "3"


def test_missing_diagnostic_line_leaves_cursor_untouched(ws_db, isolated_paths, monkeypatch):
    """No JASPER_DIAG line at all (e.g. an older outlook_scan.ps1, or a
    catastrophic COM failure before the diagnostic could even print) must
    not crash the run or write a misleading cursor value."""
    def fake_run(*a, **kw):
        return _FakeCompletedProcess("", stderr="some unrelated stderr text\n")

    monkeypatch.setattr(subprocess, "run", fake_run)
    result = oci.run(folder="Careful")

    assert result["outlook_cold_started"] is None
    assert ws_db.get_cursor("outlook_mail", "last_scan_outlook_cold_started") is None


# --- body-capture-failure diagnostic (fixed 2026-08-05, real live gap:
# raw_ref was NULL for 100% of the corpus because Save-FullBody's catch
# blocks silently swallowed every COM failure) ------------------------------

def test_parse_body_capture_failures_extracts_error_reason():
    stderr = (
        "JASPER_DIAG: outlook_was_running=True\n"
        "JASPER_DIAG: body_capture_failed field=body error=COM object failed\n"
        "some unrelated line\n"
        "JASPER_DIAG: body_capture_failed field=htmlbody error=another COM error\n"
    )
    assert oci._parse_body_capture_failures(stderr) == [
        "COM object failed",
        "another COM error",
    ]


def test_parse_body_capture_failures_extracts_no_item_dir_reason():
    stderr = "JASPER_DIAG: body_capture_failed reason=no_item_dir\n"
    assert oci._parse_body_capture_failures(stderr) == ["no_item_dir"]


def test_parse_body_capture_failures_empty_when_no_failures():
    stderr = "JASPER_DIAG: outlook_was_running=False\n"
    assert oci._parse_body_capture_failures(stderr) == []


def test_run_persists_body_capture_failure_cursors(ws_db, isolated_paths, monkeypatch):
    def fake_run(*a, **kw):
        return _FakeCompletedProcess(
            "", stderr="JASPER_DIAG: body_capture_failed field=body error=boom\n"
        )

    monkeypatch.setattr(subprocess, "run", fake_run)
    result = oci.run(folder="Careful")

    assert result["body_capture_failures"] == ["boom"]
    assert ws_db.get_cursor("outlook_mail", "body_capture_failures_total") == "1"
    assert ws_db.get_cursor("outlook_mail", "last_body_capture_failure") == "boom"


def test_run_body_capture_failure_total_accumulates_across_runs(ws_db, isolated_paths, monkeypatch):
    def fake_run(*a, **kw):
        return _FakeCompletedProcess(
            "", stderr="JASPER_DIAG: body_capture_failed field=body error=boom\n"
        )

    monkeypatch.setattr(subprocess, "run", fake_run)
    oci.run(folder="Careful")
    oci.run(folder="Careful")
    result = oci.run(folder="Careful")

    assert result["body_capture_failures"] == ["boom"]
    assert ws_db.get_cursor("outlook_mail", "body_capture_failures_total") == "3"


def test_run_no_body_capture_failures_leaves_cursors_untouched(ws_db, isolated_paths, monkeypatch):
    def fake_run(*a, **kw):
        return _FakeCompletedProcess("", stderr="JASPER_DIAG: outlook_was_running=True\n")

    monkeypatch.setattr(subprocess, "run", fake_run)
    result = oci.run(folder="Careful")

    assert result["body_capture_failures"] == []
    assert ws_db.get_cursor("outlook_mail", "body_capture_failures_total") is None


def test_sweep_unread_also_persists_body_capture_failure_cursors(ws_db, isolated_paths, monkeypatch):
    def fake_run(*a, **kw):
        return _FakeCompletedProcess(
            "", stderr="JASPER_DIAG: body_capture_failed field=htmlbody error=sweep boom\n"
        )

    monkeypatch.setattr(subprocess, "run", fake_run)
    result = oci.sweep_unread(folder="Careful")

    assert result["body_capture_failures"] == ["sweep boom"]
    assert ws_db.get_cursor("outlook_mail", "body_capture_failures_total") == "1"
    assert ws_db.get_cursor("outlook_mail", "last_body_capture_failure") == "sweep boom"


# --- timeout salvage (2026-08-05, real need: sizing a 90-day/2,849-item --
# manual backfill exposed a timeout with NO except around it at all) -----

def test_run_salvages_partial_output_on_timeout(ws_db, isolated_paths, monkeypatch, tmp_path):
    item, staged_dir = _stage_item(tmp_path, "to", "entryid-TO", "body", "<p>b</p>")

    def fake_run(*a, **kw):
        # encoding="utf-8" is passed at the real call site, so subprocess
        # itself hands TimeoutExpired a real str (not bytes) here - same
        # shape this test reproduces.
        raise subprocess.TimeoutExpired(cmd=a[0] if a else "claude", timeout=kw.get("timeout", 120),
                                         output=json.dumps(item) + "\n", stderr="")

    monkeypatch.setattr(subprocess, "run", fake_run)
    result = oci.run(folder="Careful", timeout=5)

    assert result["inserted"] == 1  # the one already-valid JSON line before the timeout is salvaged
    assert result["ok"] is False  # still reported as a real failure, not silently swallowed


def test_run_accepts_a_custom_timeout_value(ws_db, isolated_paths, monkeypatch):
    captured = {}

    def fake_run(*a, **kw):
        captured["timeout"] = kw.get("timeout")
        return _FakeCompletedProcess("")

    monkeypatch.setattr(subprocess, "run", fake_run)
    oci.run(folder="Careful", timeout=1800)

    assert captured["timeout"] == 1800


def test_schema_migration_entry_id_column_idempotent(ws_db):
    """init_workgraph()'s ALTER TABLE ADD COLUMN entry_id must be safe to run
    against an already-migrated DB (every real wake calls init_workgraph()
    again, it doesn't run once ever) - the try/except OperationalError pattern
    already used for signal_type/pr_number, applied the same way here."""
    ws_db.init_workgraph()  # second call, same db - must not raise
    ws_db.init_workgraph()  # third call for good measure
    cols = {r["name"] for r in ws_db._connect().execute("PRAGMA table_info(raw_items)").fetchall()}
    assert "entry_id" in cols
    assert "raw_ref" in cols
